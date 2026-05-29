import os
import re
import pandas as pd
from openpyxl.styles import Alignment, PatternFill, Border, Side

def process_interactive():
    user_input = input("请输入要提取的txt文件名（多个文件请用逗号隔开）：\n> ")
    user_input = user_input.replace('，', ',')
    filenames = [f.strip() for f in user_input.split(',')]
    
    all_dataframes = []
    
    for filename in filenames:
        if not filename:
            continue
            
        if not os.path.exists(filename):
            print(f"⚠️ 找不到文件 '{filename}'，已跳过。")
            continue
            
        date_match = re.search(r'物流费用提取结果(.*?)\.txt', filename)
        if not date_match:
            print(f"⚠️ 无法从 '{filename}' 提取日期作为表头，请检查命名格式。")
            continue
            
        date_col = date_match.group(1)
        
        # 兼容不同编码
        content = ""
        for enc in ['utf-8', 'gbk', 'gb2312']:
            try:
                with open(filename, 'r', encoding=enc) as f:
                    content = f.read()
                break
            except UnicodeDecodeError:
                continue
                
        # 以 "PO-" 作为订单的分割符
        orders = re.split(r'PO-', content)
        orders = [o for o in orders if o.strip()]
        
        data = []
        for order in orders:
            order_text = order.replace('\n', ' ')
            
            # 正则匹配仓库名和费用
            city_match = re.search(r'天猫美团(.*?)仓', order_text)
            fees_match = re.search(r'快递\s*([\d.]+)\s*物流\s*([\d.]+)', order_text)
            
            if city_match and fees_match:
                city = city_match.group(1).strip()
                express = float(fees_match.group(1))
                logistics = float(fees_match.group(2))
                
                data.append({'仓库': city, '货运方式': '快递', date_col: express})
                data.append({'仓库': city, '货运方式': '物流', date_col: logistics})
                
        if data:
            df = pd.DataFrame(data)
            df.set_index(['仓库', '货运方式'], inplace=True)
            all_dataframes.append(df)
            print(f"✅ {filename}: 成功提取 {len(data)//2} 个仓位的数据。")
        else:
            print(f"⚠️ {filename}: 未能提取到任何数据！")
            
    if not all_dataframes:
        print("\n❌ 没有提取到任何有效数据，程序退出！")
        return

    # 横向合并所有表
    final_df = pd.concat(all_dataframes, axis=1, join='outer')
    
    # 按照列名数字大小进行升序排序
    try:
        sorted_cols = sorted(final_df.columns, key=lambda x: float(x))
        final_df = final_df[sorted_cols]
    except ValueError:
        sorted_cols = sorted(final_df.columns)
        final_df = final_df[sorted_cols]

    # 计算总费用列
    final_df['总费用'] = final_df.sum(axis=1, skipna=True)

    # 导出 Excel 
    output_excel = "动态物流费用合并结果.xlsx"
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        final_df.to_excel(writer, sheet_name='费用汇总')
        worksheet = writer.sheets['费用汇总']
        
        # 定义样式：黄色填充、细黑边框
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        
        # 获取最大列数，用来判断哪些列不要上色 (跳过A列和最后一列)
        max_col_idx = worksheet.max_column
        
        # 遍历所有写入了数据的行和单元格
        for row in worksheet.iter_rows():
            is_express_row = False
            
            # 检查当前行是否是“快递”行（B列，即索引1）
            if len(row) > 1 and row[1].value == '快递':
                is_express_row = True
                
            for cell in row:
                # 统一加上框线和居中
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 精准上色：如果是快递行，且不是A列(仓库，idx=1)，且不是最后一列(总费用)才上色
                # openpyxl 的 column 从 1 开始算
                if is_express_row and cell.column > 1 and cell.column < max_col_idx:
                    cell.fill = yellow_fill

        # 调整所有列的列宽
        for col in worksheet.columns:
            column = col[0].column_letter
            worksheet.column_dimensions[column].width = 14
                
    print(f"\n🎉 处理完成！已增加框线并优化填充范围，文件保存为: {output_excel}")

if __name__ == "__main__":
    process_interactive()
